from __future__ import annotations

import json
import os
import shutil
import tempfile
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import (
    PMF_ACTIVE_PATH,
    PMF_META_PATH,
    RAW_MATERIAL_SHEET,
)
from app.services.pmf_service import read_pmf_bundle
from app.services.supplier_service import clean, get_full_row_data


@dataclass(frozen=True)
class PmfMaterialSnapshot:
    row_pos: int
    depth: int
    material_no: str
    supplier: str
    material_name: str
    english_name: str
    maker: str
    maker_country: str
    org: str
    cert_no: str
    expiry_date: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class PmfUpdateResult:
    ok: bool
    pmf_path: str
    backup_path: str
    row_pos: int
    excel_row: int
    depth: int
    changed_fields: dict[str, dict[str, str]]
    updated_at: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def normalize_material_no(value: Any) -> str:
    text = clean(value)
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in {"-", "nan", "none", "nat"}:
        return ""

    try:
        from pandas import to_datetime

        parsed = to_datetime(text, errors="coerce")
        if parsed is None or str(parsed) == "NaT":
            return ""
        return parsed.strftime("%Y-%m-%d")
    except Exception:
        return ""


def get_pmf_material_snapshot(
    row_pos: int,
    depth: int = 0,
) -> PmfMaterialSnapshot:
    pmf_path = resolve_pmf_update_path()

    active_path = Path(
        PMF_ACTIVE_PATH
    )

    if (
        pmf_path.resolve()
        == active_path.resolve()
    ):
        bundle = read_pmf_bundle()
        df_raw = bundle["df_raw"]

    else:
        if not pmf_path.exists():
            raise FileNotFoundError(
                "PMF update file not found: "
                + str(pmf_path)
            )

        import pandas as pd

        keep_vba = (
            pmf_path.suffix.lower()
            == ".xlsm"
        )

        workbook = load_workbook(
            pmf_path,
            read_only=True,
            data_only=True,
            keep_vba=keep_vba,
        )

        try:
            sheet_name = _find_sheet_name(
                workbook,
                RAW_MATERIAL_SHEET,
            )
        finally:
            workbook.close()

        df_raw = pd.read_excel(
            pmf_path,
            sheet_name=sheet_name,
            header=None,
            engine="openpyxl",
        )

    if (
        row_pos < 0
        or row_pos >= len(df_raw)
    ):
        raise ValueError(
            "PMF row_pos out of range: "
            + str(row_pos)
        )

    if depth < 0 or depth > 4:
        raise ValueError(
            "PMF depth out of range: "
            + str(depth)
        )

    row = df_raw.iloc[row_pos]

    main = (
        get_full_row_data(row, 0)
        or {}
    )

    selected = (
        get_full_row_data(row, depth)
        or {}
    )

    material_name = clean(
        selected.get("n")
    )

    if material_name in {"", "-"}:
        raise ValueError(
            "No valid material at PMF row: "
            + str(row_pos)
            + ", depth: "
            + str(depth)
        )

    return PmfMaterialSnapshot(
        row_pos=int(row_pos),
        depth=int(depth),
        material_no=normalize_material_no(
            main.get("id")
        ),
        supplier=(
            clean(row.iloc[6])
            if len(row) > 6
            else "-"
        ),
        material_name=material_name,
        english_name=clean(
            selected.get("e")
        ),
        maker=clean(
            selected.get("m")
        ),
        maker_country=clean(
            selected.get("o")
        ),
        org=clean(
            selected.get("h")
        ),
        cert_no=clean(
            selected.get("i")
        ),
        expiry_date=normalize_date(
            selected.get("v")
        ),
    )



def get_pmf_update_columns(depth: int) -> dict[str, int]:
    """
    openpyxl 1-based 컬럼 번호.
    메인원료: H/I/J
    하부원료: O열(원료명)부터 9열 단위 반복, org/cert/expiry는 +4/+5/+6.
    """
    if depth == 0:
        return {"org": 8, "cert_no": 9, "expiry_date": 10}

    if depth < 1 or depth > 4:
        raise ValueError(f"지원하지 않는 PMF depth입니다: {depth}")

    base_zero = 15 + (depth - 1) * 9
    return {
        "org": base_zero + 5,
        "cert_no": base_zero + 6,
        "expiry_date": base_zero + 7,
    }


def resolve_pmf_update_path() -> Path:
    """
    테스트 중에는 PMF_UPDATE_PATH 환경변수로 별도 xlsm 복사본을 지정할 수 있다.
    미지정 시 active_pmf.xlsm 캐시만 업데이트한다. 네트워크 원본은 직접 수정하지 않는다.
    """
    configured = os.getenv("PMF_UPDATE_PATH", "").strip()
    return Path(configured) if configured else Path(PMF_ACTIVE_PATH)


def _find_sheet_name(workbook, target: str) -> str:
    if target in workbook.sheetnames:
        return target

    target_norm = target.strip().lower()
    for name in workbook.sheetnames:
        if str(name).strip().lower() == target_norm:
            return name

    raise ValueError(f"PMF 시트를 찾지 못했습니다: {target}")


def _display_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    text = str(value).strip()
    return "" if text.lower() in {"", "-", "nan", "none", "nat"} else text


def preview_pmf_update(
    row_pos: int,
    depth: int,
    cert_org: str,
    cert_no: str = "",
    expiry_date: str = "",
) -> dict[str, Any]:
    snapshot = get_pmf_material_snapshot(row_pos=row_pos, depth=depth)
    new_org = str(cert_org or "").strip().upper()
    new_cert_no = str(cert_no or "").strip()
    new_expiry = normalize_date(expiry_date)

    changes = {
        "org": {"before": "" if snapshot.org == "-" else snapshot.org, "after": new_org},
        "cert_no": {
            "before": "" if snapshot.cert_no == "-" else snapshot.cert_no,
            "after": new_cert_no,
        },
        "expiry_date": {"before": snapshot.expiry_date, "after": new_expiry},
    }

    warnings: list[str] = []
    if not new_org:
        warnings.append("인증기관이 없습니다.")

    if new_expiry and snapshot.expiry_date and new_expiry < snapshot.expiry_date:
        warnings.append(
            f"유효기간이 기존값보다 과거입니다: {snapshot.expiry_date} -> {new_expiry}"
        )

    return {
        "snapshot": snapshot.to_dict(),
        "pmf_path": str(resolve_pmf_update_path()),
        "changes": changes,
        "warnings": warnings,
    }


def update_pmf_certificate_fields(
    row_pos: int,
    depth: int,
    cert_org: str,
    cert_no: str = "",
    expiry_date: str = "",
    allow_date_regression: bool = False,
) -> PmfUpdateResult:
    preview = preview_pmf_update(
        row_pos=row_pos,
        depth=depth,
        cert_org=cert_org,
        cert_no=cert_no,
        expiry_date=expiry_date,
    )

    if preview["warnings"] and not allow_date_regression:
        date_warning = [x for x in preview["warnings"] if "과거" in x]
        if date_warning:
            raise ValueError(" / ".join(date_warning))

    pmf_path = resolve_pmf_update_path()
    if not pmf_path.exists():
        raise FileNotFoundError(f"업데이트 대상 PMF 파일을 찾지 못했습니다: {pmf_path}")

    backup_dir = pmf_path.parent / "pmf_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = backup_dir / f"{pmf_path.stem}_{stamp}{pmf_path.suffix}"
    shutil.copy2(pmf_path, backup_path)

    keep_vba = pmf_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(pmf_path, keep_vba=keep_vba)
    sheet_name = _find_sheet_name(workbook, RAW_MATERIAL_SHEET)
    worksheet = workbook[sheet_name]

    excel_row = int(row_pos) + 1
    columns = get_pmf_update_columns(depth)

    changed_fields: dict[str, dict[str, str]] = {}
    updates = {
        "org": str(cert_org or "").strip().upper(),
        "cert_no": str(cert_no or "").strip(),
        "expiry_date": normalize_date(expiry_date),
    }

    for field, new_value in updates.items():
        # 빈값으로 기존 PMF 값을 지우지 않는다.
        if not new_value:
            continue

        cell = worksheet.cell(row=excel_row, column=columns[field])
        before = _display_cell_value(cell.value)

        if field == "expiry_date":
            parsed = datetime.strptime(new_value, "%Y-%m-%d")
            cell.value = parsed
            cell.number_format = "yyyy-mm-dd"
        else:
            cell.value = new_value

        after = new_value
        if before != after:
            changed_fields[field] = {"before": before, "after": after}

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{pmf_path.stem}_",
            suffix=pmf_path.suffix,
            dir=pmf_path.parent,
            delete=False,
        ) as temp_file:
            temp_path = Path(temp_file.name)

        workbook.save(temp_path)
        workbook.close()
        os.replace(temp_path, pmf_path)
        temp_path = None
    except Exception:
        try:
            workbook.close()
        except Exception:
            pass
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)
        shutil.copy2(backup_path, pmf_path)
        raise

    if Path(PMF_META_PATH).exists() and pmf_path == Path(PMF_ACTIVE_PATH):
        try:
            meta = json.loads(Path(PMF_META_PATH).read_text(encoding="utf-8"))
        except Exception:
            meta = {}
        meta["active_pmf_updated_at"] = datetime.now().isoformat(timespec="seconds")
        meta["active_pmf_update_reason"] = "certificate_filing_confirm"
        Path(PMF_META_PATH).write_text(
            json.dumps(meta, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return PmfUpdateResult(
        ok=True,
        pmf_path=str(pmf_path),
        backup_path=str(backup_path),
        row_pos=int(row_pos),
        excel_row=excel_row,
        depth=int(depth),
        changed_fields=changed_fields,
        updated_at=datetime.now().isoformat(timespec="seconds"),
    )


def restore_pmf_backup(backup_path: str | Path, pmf_path: str | Path) -> None:
    source = Path(backup_path)
    target = Path(pmf_path)
    if not source.exists():
        raise FileNotFoundError(f"PMF 백업 파일을 찾지 못했습니다: {source}")
    shutil.copy2(source, target)
